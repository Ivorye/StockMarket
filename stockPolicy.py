import tushare as ts
import openpyxl
from pandas import DataFrame
import time
import datetime
import os
import csv
import mysql.connector

TUSHARE_TOKEN = '4d47c02a8bb025881c9dd9e3c36d25139ab5b429a73353e566fc02a9'

def _check_execution_status(task_name):
	"""检查任务是否已成功执行过"""
	try:
		mdb=_connect_sm()
		mycsr=mdb.cursor()
		mycsr.execute("SELECT status FROM st_execution_status WHERE task_name=%s",(task_name,))
		row=mycsr.fetchone()
		mycsr.close();mdb.close()
		return row and row[0]=='Y'
	except Exception:
		return False

def _update_execution_status(task_name, status='Y'):
	"""更新任务执行状态"""
	try:
		mdb=_connect_sm()
		mycsr=mdb.cursor()
		today=datetime.date.today().strftime('%Y%m%d')
		sql="INSERT INTO st_execution_status(task_name,exec_date,status) VALUES(%s,%s,%s) ON DUPLICATE KEY UPDATE exec_date=%s,status=%s"
		mycsr.execute(sql,(task_name,today,status,today,status))
		mdb.commit()
		mycsr.close();mdb.close()
	except Exception as e:
		print(f'更新执行状态失败: {e}')

#延迟加载：优先从stocks表读取，回退到tushare API
_basic_cache = None
def _get_stock_basic():
	global _basic_cache
	if _basic_cache is not None:
		return _basic_cache
	#先从数据库读取
	try:
		mdb=_connect_sm()
		mycsr=mdb.cursor()
		mycsr.execute("SELECT COUNT(*) FROM stocks")
		cnt=mycsr.fetchone()[0]
		if cnt>0:
			mycsr.execute("SELECT st_code AS ts_code,symbol,fullname AS name,list_date FROM stocks")
			rows=mycsr.fetchall()
			df=DataFrame(rows,columns=['ts_code','symbol','name','list_date'])
			_basic_cache=df
			mycsr.close();mdb.close()
			return _basic_cache
		mycsr.close();mdb.close()
	except Exception:
		pass
	#回退到tushare API
	try:
		pro = ts.pro_api(TUSHARE_TOKEN)
		_basic_cache = pro.query('stock_basic')
		#缓存到stocks表
		_store_to_stocks(_basic_cache)
		_update_execution_status('stock_basic','Y')
		return _basic_cache
	except Exception as e:
		if _check_execution_status('stock_basic'):
			print('tushare API失败，st_execution_status状态为Y，从stocks表重试读取')
			mdb=_connect_sm()
			mycsr=mdb.cursor()
			mycsr.execute("SELECT st_code AS ts_code,symbol,fullname AS name,list_date FROM stocks")
			rows=mycsr.fetchall()
			mycsr.close();mdb.close()
			if rows:
				_basic_cache=DataFrame(rows,columns=['ts_code','symbol','name','list_date'])
				return _basic_cache
		print('tushare API失败且无成功执行记录: %s'%e)
		raise

def _store_to_stocks(df):
	mdb=_connect_sm()
	mycsr=mdb.cursor()
	sql="INSERT IGNORE INTO stocks(id,symbol,st_code,fullname,list_date) VALUES(%s,%s,%s,%s,%s)"
	for i in range(len(df)):
		try:
			mycsr.execute(sql,(i+1,df.iloc[i].symbol,df.iloc[i].ts_code,df.iloc[i].name,df.iloc[i].list_date))
		except Exception:
			pass
		if i%500==499:
			mdb.commit()
	mdb.commit()
	mycsr.close()
	mdb.close()

# 从 st_daily 查询日线数据，返回与 ts.pro_bar 相同结构的 DataFrame
#columns: trade_date, open, high, low, close, pct_chg, vol
def _get_daily_from_db(symbol, startDate, endDate):
	mdb=_connect_sm()
	mycsr=mdb.cursor()
	sql="SELECT trade_date,openp,high,low,closep,pct_chg,vol FROM st_daily WHERE symbol=%s AND trade_date BETWEEN %s AND %s ORDER BY trade_date DESC"
	mycsr.execute(sql,(symbol,startDate,endDate))
	rows=mycsr.fetchall()
	mycsr.close()
	mdb.close()
	if not rows:
		return None
	return DataFrame(rows,columns=['trade_date','open','high','low','close','pct_chg','vol'])

def _connect_sm():
	return mysql.connector.connect(host="localhost",user="root",passwd="P@ssw0rd",database='stockshare',connection_timeout=10)

#通用筛选条件：剔除科创板(688)、ST、*ST、次新股(2021年后上市且上市晚于startDate)
def _should_skip(df, i, startDate=''):
	sym = df.symbol[i]
	name = df.name[i]
	lst = df.list_date[i]
	if sym[0:3] == '688':
		return True
	if name[0:2] == 'ST' or name[0:2] == '*S':
		return True
	#原逻辑：list_date < '20210101' AND list_date < startDate 两个条件同时满足才不排除
	if lst >= '20210101' and (not startDate or lst >= startDate):
		return True
	return False

#获取默认股票列表
def _default_stock_list():
	df = _get_stock_basic()
	return df.ts_code

#def myThread(threadName='', param=''):

def loadAllStockBasicsIntoMysql():
	pro = ts.pro_api(TUSHARE_TOKEN)
	df  = pro.query('stock_basic')
	stockList = df.ts_code
	l = len(stockList)
	mdb=_connect_sm()
	mycsr = mdb.cursor()
	sql = "insert into st_basic(ts_code,symbol,name,area,list_date) values(%s,%s,%s,%s,%s)"
	sql2= "select * from st_basic where ts_code=%s"
	counter=0
	print(time.ctime(),'-------- processing begin---------------')
#load all basic table data
	for i in range(len(df)):
		val2= (df.ts_code[i],)
		mycsr.execute(sql2,val2)
		rst=mycsr.fetchall()
		if(len(rst) ==0):
			tscode= df.ts_code[i]
			symbol= df.symbol[i]
			name  = df.name[i]
			area  = df.area[i]
			indst = df.industry[i]
			market= df.market[i]
			lstdte= df.list_date[i]
			val =(tscode,symbol,name,area,lstdte)
			mycsr.execute(sql,val)
			counter+=1
		if(i%100 == 99):
			print(round(i/100)*100, 'records processed...')
		if(i== len(df) -1):
			print('All', len(df), 'records processed! added ',counter,' new records')
	mdb.commit()
	mycsr.close()
	mdb.close()
	return stockList

#个股检测上涨动能算法：一个月内出现过单日6个点上涨的幅度(A日)。月底较月初上涨，A日过后振幅收窄且量缩（所有交易日的单日量没有一天超过A日总量），
#一个星期后可以买入，或等再一次涨幅超3个点时介入。

#选出迄今跳空过
def gettiaokongshangzhangguo(startDate='',endDate=''):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return []
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime())
		print('end date is:',endDate)
	df = _get_stock_basic()
	stockList = df.ts_code
	l = len(stockList)
	lstMultiple = []
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
		if (not _should_skip(df, i, startDate)):
			hangqing = _get_daily_from_db(df.symbol[i], startDate, endDate)
			if hangqing is None:
				continue
			idx = len(hangqing)
			if (idx > 3):
				x = 0
				for j in range(idx-1):
					volChange = ( hangqing.vol[j] / hangqing.vol[j + 1] )
					if (hangqing.close[0]>hangqing.close[idx-1]*1.3
					and (hangqing.low[j]>hangqing.high[j+1]
					or (hangqing.open[j+1]<hangqing.close[j+1]<hangqing.open[j]
					and volChange>2.5))):
						x = 1
						break
				if (x == 1):
					lstMultiple.append(df.ts_code[i])
					print(df.ts_code[i],j,hangqing.vol[j+1], hangqing.vol[j], round(volChange))
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")
	print(time.ctime(),'-------- processing end-----------------')
	if lstMultiple:
		_write_signal(lstMultiple, '跳空上涨过', endDate)
	return lstMultiple

#选出曾经有过巨量上涨记录，第二三天量跌调整，但是振幅在2%以内的股票，------->很有效，选中股票莱美药业买入直赚10个点
def getJuliangshangzhang(stockList='',startDate='',endDate='',multiple=''):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return []
	if(multiple is None or multiple == ''):
		multiple = 2;print('multiple is ', multiple)
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime())
	if(stockList is None or stockList == ''):
		stockList = _default_stock_list()
	df = _get_stock_basic()
	l = len(stockList)
	lstMultiple = []
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
		if (not _should_skip(df, i, startDate)):
			hangqing = _get_daily_from_db(df.symbol[i], startDate, endDate)
			if hangqing is None:
				continue
			idx = len(hangqing)
			if (idx > 3):
				x = 0
				for j in range(idx-1):
					volChange = ( hangqing.vol[j] / hangqing.vol[j + 1] )
					if (volChange >= multiple and j > 2
					and hangqing.open[j]>hangqing.close[j+1]
					and hangqing.pct_chg[j]>4
					and ( hangqing.close[j-1]>hangqing.close[j] and -3 < hangqing.pct_chg[j-1]<3 and hangqing.vol[j] >= hangqing.vol[j-1]*multiple
					or hangqing.close[j-2]> hangqing.close[j] and -3 <hangqing.pct_chg[j-2]<3 and hangqing.vol[j] >= hangqing.vol[j-2]*multiple)
#or hangqing.change[j-2] < 3
					and hangqing.close[0] > hangqing.close[idx-1]):
						x = 1
						break
				if (x == 1):
					lstMultiple.append(df.ts_code[i])
					print(df.ts_code[i],j,hangqing.vol[j+1], hangqing.vol[j], round(volChange))
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")
	print(time.ctime(),'-------- processing end-----------------')
	if lstMultiple:
		_write_signal(lstMultiple, '巨量上涨', endDate)
	return lstMultiple

#获取向上跳空的股票，剔除科创版、ST和次新股
def getxiangshangtiaokongquekou(stockList='',startDate='',endDate=''):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return []
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime());print('endDate is ', endDate)
	if(stockList is None or stockList == ''):
		stockList = _default_stock_list()
	df = _get_stock_basic()
	l = len(stockList)
	lstMultiple = []
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
		if (not _should_skip(df, i, startDate)):
			hangqing = _get_daily_from_db(df.symbol[i], startDate, endDate)
			if hangqing is None:
				continue
			idx = len(hangqing)
			if (idx > 3):
				x = 0
				volChange = ( hangqing.vol[0] / hangqing.vol[1] )
				if (hangqing.low[0] > hangqing.high[1]
				and hangqing.close[0] > hangqing.close[idx-1]):
					x = 1
				if (x == 1):
					lstMultiple.append(df.ts_code[i])
					print(df.ts_code[i],hangqing.vol[1], hangqing.vol[0], round(volChange))
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")
	print(time.ctime(),'-------- processing end-----------------')
	if lstMultiple:
		_write_signal(lstMultiple, '向上跳空缺口', endDate)
	return lstMultiple

def getStockListByFluxRate(startDate='',endDate='',fluxRate='',filePath=''):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime())
	if(fluxRate is None or fluxRate == ''):
		fluxRate = 50.00
	df = _get_stock_basic()
	stockList = df.ts_code
	l = len(stockList)
	fluxs={}
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
		if (not _should_skip(df, i)):
			hangqing = _get_daily_from_db(df.symbol[i], startDate, endDate)
			if (hangqing is not None):
				idx = len(hangqing)
				if (idx > 1):
					flux = ( hangqing.close[0] - hangqing.close[idx-1] ) / hangqing.close[idx-1] * 100
					flux2 = round(flux,2)
					if (flux2 >= fluxRate):
						fluxs[df.ts_code[i]] = flux2
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")


	print(time.ctime(),'-------- processing ended---------------')
	flxs=sorted(fluxs.items(),key=lambda x:x[1],reverse = True)
	resultDf = DataFrame(flxs)
	resultDf.to_excel(filePath)

#选出曾经有过巨量上涨记录的股票，
def getStockListByVolumeChange(stockList='',startDate='',endDate='',multiple=''):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return []
	if(multiple is None or multiple == ''):
		print('multiple must input')
		return []
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime())
	if(stockList is None or stockList == ''):
		stockList = _default_stock_list()
	l = len(stockList)
	lstStocks=[]
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
#		if (df.symbol[i][0:3]!= '688' and df.name[i][0:2]!='ST' and df.name[i][0:2] !='*S' and df.list_date[i]<'20210101'
#		and df.list_date[i] < startDate):
		hangqing = _get_daily_from_db(stockList[i].split('.')[0], startDate, endDate)
		if(hangqing is not None):
			idx = len(hangqing)
			if (idx > 1):
				x = 0
				for j in range(idx-1):
					volChange = ( hangqing.vol[j] / hangqing.vol[j + 1] )
					if (volChange >= multiple and hangqing.open[j]>hangqing.close[j+1] and (hangqing.close[j]-hangqing.close[j+1])>3
					and hangqing.close[0] > hangqing.close[idx-1]):
						x = 1
						break
				if (x == 1):
					lstStocks.append(stockList[i])
		# print(df.ts_code[i],j,hangqing.vol[j+1], hangqing.vol[j], round(volChange))
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")
	if lstStocks:
		_write_signal(lstStocks, '放量变化', endDate)
	return lstStocks

def getLiangzeng(stockList='',startDate='',endDate='',multiple=''):
	if(startDate is None or startDate == ''):
		print('startDate must input'); return []
	if(multiple is None or multiple == ''):
		print('multiple must input');return []
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime())
	if(stockList is None or stockList == ''):
		print('stockList must input');return []
	l = len(stockList)
	lstStocks=[]
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
		hangqing = _get_daily_from_db(stockList[i].split('.')[0], startDate, endDate)
		if(hangqing is not None):
			idx = len(hangqing)
			if (idx >= 9):
				x = 0
				for j in range(5):
					if(hangqing.vol[j + 1] > hangqing.vol[0] * multiple
					and (hangqing.vol[j + 1] > hangqing.vol[j+2] * 2.5 or hangqing.vol[j + 1] > hangqing.vol[j+3] * 2.5 or hangqing.vol[j + 1] > hangqing.vol[j+4] * 2.5)
					and hangqing.pct_chg [j+1] > 4
					and hangqing.close[0] > hangqing.close[idx-1]
					and hangqing.close[0] > hangqing.close[j+1]):
						x = 1; break
				if (x == 1):
					lstStocks.append(stockList[i])
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")
	if lstStocks:
		_write_signal(lstStocks, '量增', endDate)
	return lstStocks

#获取放量股票，当日放量倍数由multiple确定
def getFangliangDay0(stockList='',startDate='',endDate='',multiple=''):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return []
	if(multiple is None or multiple == ''):
		print('multiple must input')
		return []
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d",time.localtime())
	if(stockList is None or stockList == ''):
		stockList = _default_stock_list()
	l = len(stockList)
	lstMultiple = []
	print(time.ctime(),'-------- processing begin---------------')
	for i in range(l):
		hangqing = _get_daily_from_db(stockList[i].split('.')[0], startDate, endDate)
		if(hangqing is not None):
			idx = len(hangqing)
			if (idx > 1):
				if (hangqing.vol[0]> hangqing.vol[1] * multiple):
					lstMultiple.append(stockList[i])
		if ( i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if ( i== l -1):
			print (time.ctime(), " All records have been processed!!!")
	print(time.ctime(),'-------- processing end-----------------')
	if lstMultiple:
		_write_signal(lstMultiple, '放量日', endDate)
	return lstMultiple


#筛选区间内涨幅超过指定百分比的股票，排除科创板/ST/次新股
def getZhangFu(startDate='', endDate='', pct=30):
	if(startDate is None or startDate == ''):
		print('startDate must input')
		return []
	if(endDate is None or endDate == ''):
		endDate = time.strftime("%Y%m%d", time.localtime())
	df = _get_stock_basic()
	stockList = df.ts_code
	l = len(stockList)
	lst = []
	print(time.ctime(), '-------- processing begin---------------')
	for i in range(l):
		if(not _should_skip(df, i, startDate)):
			hangqing = _get_daily_from_db(df.symbol[i], startDate, endDate)
			if hangqing is not None and len(hangqing) > 1:
				idx = len(hangqing)
				change = round((hangqing.close[0] - hangqing.close[idx-1]) / hangqing.close[idx-1] * 100, 2)
				if change >= pct:
					lst.append((df.ts_code[i], change))
					print(df.ts_code[i], f'{change}%')
		if(i % 100 == 99):
			print(time.ctime(), round(i/100)*100, ' records have been processed....')
		if(i == l - 1):
			print(time.ctime(), " All records have been processed!!!")
	print(time.ctime(), '-------- processing end-----------------')
	lst.sort(key=lambda x: x[1], reverse=True)
	codes = [item[0] for item in lst]
	if codes:
		_write_signal(codes, f'区间涨幅{pct}%', endDate)
	return codes


#6个点代表强势，相对于T-1或T-2日放量2倍以上且涨超6个点，寻找这样的股票。西藏珠峰7.2日启动，中泰股份20210826晚关注到

#获取Excel里的股票list，返回一个list出去
def getlistgupiao(file):
	if(file is None or file == ''):
		print('file must be input');return
	lista=[]
	wb=openpyxl.load_workbook(filename=file)
	sht=wb['Sheet1']
	for i in range(sht.max_row):
		lista.append(sht.cell(i+1,1).value)
	return lista


#========== 放量涨幅筛选（st_daily_signal）==========

#创建 st_daily 表（汇总所有个股日线数据）
def createDailyTable():
	mdb=_connect_sm()
	mycsr=mdb.cursor()
	sql="CREATE TABLE IF NOT EXISTS st_daily (" \
		"ts_code VARCHAR(12) NOT NULL," \
		"symbol VARCHAR(6) NOT NULL," \
		"trade_date VARCHAR(8) NOT NULL," \
		"openp FLOAT," \
		"high FLOAT," \
		"low FLOAT," \
		"closep FLOAT," \
		"preclose FLOAT," \
		"changes FLOAT," \
		"pct_chg FLOAT," \
		"vol FLOAT," \
		"amount FLOAT," \
		"PRIMARY KEY (ts_code, trade_date)," \
		"INDEX idx_trade_date (trade_date)," \
		"INDEX idx_symbol (symbol)" \
		")"
	mycsr.execute(sql)
	mdb.commit()
	mycsr.close()
	mdb.close()
	print('st_daily table created')

# 舊排程相容入口；日線載入現在直接寫入 st_daily。
def buildDailyTable(startDate='', endDate=''):
	createDailyTable()
	print(time.ctime(),'st_daily is the primary daily-data table; no aggregation needed')

#创建 st_daily_signal 表（策略信号表）
def createSignalTable():
	mdb=_connect_sm()
	mycsr=mdb.cursor()
	sql="CREATE TABLE IF NOT EXISTS st_daily_signal (" \
		"id INT AUTO_INCREMENT PRIMARY KEY," \
		"trade_date VARCHAR(8) NOT NULL," \
		"st_code VARCHAR(12) NOT NULL," \
		"strategy VARCHAR(50) NOT NULL DEFAULT 'default'," \
		"closePrice FLOAT," \
		"pct_chg FLOAT," \
		"vol FLOAT," \
		"prev_vol FLOAT," \
		"vol_ratio FLOAT," \
		"created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP," \
		"UNIQUE KEY uk_date_code_strategy (trade_date, st_code, strategy)" \
		")"
	mycsr.execute(sql)
	# 兼容旧表：添加 strategy 列并更新唯一键
	try:
		mycsr.execute("ALTER TABLE st_daily_signal ADD COLUMN strategy VARCHAR(50) NOT NULL DEFAULT 'default' AFTER st_code")
		mdb.commit()
	except Exception:
		pass
	try:
		mycsr.execute("ALTER TABLE st_daily_signal DROP INDEX uk_date_code")
		mdb.commit()
	except Exception:
		pass
	try:
		mycsr.execute("ALTER TABLE st_daily_signal ADD UNIQUE KEY uk_date_code_strategy (trade_date, st_code, strategy)")
		mdb.commit()
	except Exception:
		pass
	mycsr.close()
	mdb.close()
	print('st_daily_signal table created')

#追加数据到CSV文件，按交易日分文件，存放在output目录
def _append_csv(trade_date, rows):
	output_dir=os.path.join(os.path.dirname(os.path.abspath(__file__)),'output')
	os.makedirs(output_dir,exist_ok=True)
	filepath=os.path.join(output_dir,f'signals_{trade_date}.csv')
	write_header=not os.path.exists(filepath)
	with open(filepath,'a',newline='',encoding='utf-8-sig') as f:
		w=csv.writer(f)
		if write_header:
			w.writerow(['trade_date','st_code','strategy','closePrice','pct_chg','vol','prev_vol','vol_ratio'])
		w.writerows(rows)

#将股票列表的信号写入 st_daily_signal 表
#strategy: 策略名称，st_codes: ts_code列表，trade_date: 交易日
def _write_signal(st_codes, strategy, trade_date=''):
	mdb=_connect_sm()
	mycsr=mdb.cursor()
	if not trade_date:
		mycsr.execute("SELECT MAX(trade_date) FROM st_daily")
		r=mycsr.fetchone()
		trade_date=r[0] if r else ''
	if not trade_date:
		mycsr.close();mdb.close()
		return
	sql_insert="INSERT IGNORE INTO st_daily_signal(trade_date,st_code,strategy,closePrice,pct_chg,vol,prev_vol,vol_ratio) " \
		"SELECT a.ts_code, %s, %s, a.closep, a.pct_chg, a.vol, b.vol, ROUND(a.vol/b.vol,2) " \
		"FROM st_daily a LEFT JOIN st_daily b ON a.ts_code=b.ts_code " \
		"WHERE a.ts_code=%s AND a.trade_date=%s AND b.trade_date=(" \
		"SELECT DISTINCT trade_date FROM st_daily WHERE trade_date<%s ORDER BY trade_date DESC LIMIT 1)"
	# 查询实际写入的数据用于CSV
	sql_query="SELECT %s AS trade_date, a.ts_code, %s AS strategy, a.closep, a.pct_chg, a.vol, b.vol AS prev_vol, ROUND(a.vol/b.vol,2) AS vol_ratio " \
		"FROM st_daily a LEFT JOIN st_daily b ON a.ts_code=b.ts_code " \
		"WHERE a.ts_code=%s AND a.trade_date=%s AND b.trade_date=(" \
		"SELECT DISTINCT trade_date FROM st_daily WHERE trade_date<%s ORDER BY trade_date DESC LIMIT 1)"
	inserted=0
	csv_rows=[]
	for code in st_codes:
		mycsr.execute(sql_insert,(trade_date,strategy,code,trade_date,trade_date))
		inserted+=mycsr.rowcount
		mycsr.execute(sql_query,(trade_date,strategy,code,trade_date,trade_date))
		r=mycsr.fetchone()
		if r:
			csv_rows.append(list(r))
	mdb.commit()
	mycsr.close();mdb.close()
	print(f'_write_signal [{strategy}]: {len(st_codes)} stocks, {inserted} new records for {trade_date}')
	if csv_rows:
		_append_csv(trade_date, csv_rows)

#筛选放量涨幅股票：当日成交量>=前日3倍且涨幅>6%，排除科创板/ST/次新股
#结果写入 st_daily_signal 表（去重插入）
def getLiangJiaFangLiang(startDate='', endDate=''):
	mdb=_connect_sm()
	mycsr=mdb.cursor()
	#获取st_daily中最近两个交易日
	if(endDate is None or endDate == ''):
		mycsr.execute("SELECT DISTINCT trade_date FROM st_daily ORDER BY trade_date DESC LIMIT 2")
		dates=mycsr.fetchall()
		if(len(dates)<2):
			print('st_daily数据不足两个交易日，请先运行buildDailyTable()')
			mycsr.close();mdb.close()
			return []
		today=dates[0][0]
		prev_day=dates[1][0]
	else:
		today=endDate
		mycsr.execute("SELECT DISTINCT trade_date FROM st_daily WHERE trade_date<%s ORDER BY trade_date DESC LIMIT 1",(today,))
		r=mycsr.fetchone()
		if not r:
			print('找不到前一交易日')
			mycsr.close();mdb.close()
			return []
		prev_day=r[0]
	#SQL联查：当日成交量>=前日3倍且涨幅>6%
	sql="SELECT a.ts_code,a.trade_date,a.closep,a.pct_chg,a.vol,b.vol AS prev_vol," \
		"ROUND(a.vol/b.vol,2) AS vol_ratio " \
		"FROM st_daily a JOIN st_daily b ON a.ts_code=b.ts_code " \
		"WHERE a.trade_date=%s AND b.trade_date=%s " \
		"AND a.vol>=b.vol*3 AND a.pct_chg>6"
	mycsr.execute(sql,(today,prev_day))
	rows=mycsr.fetchall()
	#加载stock_basic用于_should_skip过滤
	df=None
	ts_to_idx={}
	try:
		df=_get_stock_basic()
		ts_to_idx={df.ts_code[i]:i for i in range(len(df))}
	except Exception:
		pass
	result=[]
	for row in rows:
		ts_code=row[0]
		sym=ts_code.split('.')[0]
		#排除科创板(688)
		if(sym.startswith('688')):
			continue
		#如果有stock_basic数据，应用完整_should_skip过滤
		if df is not None:
			idx=ts_to_idx.get(ts_code)
			if idx is not None and _should_skip(df,idx):
				continue
		result.append({
			'st_code':row[0],'trade_date':row[1],'closePrice':row[2],
			'pct_chg':row[3],'vol':row[4],'prev_vol':row[5],'vol_ratio':row[6]
		})
	#写入st_daily_signal（去重插入）
	sql_insert="INSERT IGNORE INTO st_daily_signal(trade_date,st_code,strategy,closePrice,pct_chg,vol,prev_vol,vol_ratio) " \
		"VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"
	inserted=0
	for r in result:
		mycsr.execute(sql_insert,(r['trade_date'],r['st_code'],'量价放量',r['closePrice'],r['pct_chg'],r['vol'],r['prev_vol'],r['vol_ratio']))
		inserted+=mycsr.rowcount
	mdb.commit()
	mycsr.close()
	mdb.close()
	print(time.ctime(),'getLiangJiaFangLiang: %s vs %s, 筛选出%d只, 写入%d条新记录'%(today,prev_day,len(result),inserted))
	return result

